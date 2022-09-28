import os.path
from openpyxl import load_workbook

wb = load_workbook(filename = os.path.join(os.getcwd(), 'budget.xlsx'))
ws = wb['Transactions']
print(ws['B5'].value)
wb.save(os.path.join(os.getcwd(), 'budget.xlsx'))
